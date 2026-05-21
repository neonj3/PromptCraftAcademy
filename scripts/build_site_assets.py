from __future__ import annotations

import json
import re
from collections import OrderedDict
from pathlib import Path

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"

DOC_100 = Path(r"C:\Users\ASUS\Downloads\AI_Generative_AI_100_Page_Beginner_Workbook.docx")
DOC_DETAILED = Path(r"C:\Users\ASUS\Downloads\Detailed_Beginner_AI_Workbook.docx")
DOC_REVISION = Path(r"C:\Users\ASUS\Downloads\Generative_AI_Workbook.docx")

SITE_JSON = DATA_DIR / "site-content.json"
DB_WORKBOOK = DATA_DIR / "promptcraft-github-database.xlsx"


TEACHERS = [
    {
        "teacherId": "gaurav-sen",
        "teacherName": "Gaurav Sen",
        "channel": "The Engineering Glossary",
        "domain": "Core technical concepts for builders",
        "summary": "Explains the mechanics behind tokens, vectors, attention, and context systems in language that aspiring AI builders can reuse.",
        "teachingStyle": "Builder-first, terminology-aware, practical systems thinking",
        "signatureTopics": ["LLMs", "tokenization", "attention", "RAG", "vector databases"],
        "teacherVoice": "If you are building applications, learn the language of AI so deeper subjects become easier to understand.",
        "realWorldImpact": "Ideal for developers and technical learners who want sharper engineering intuition.",
    },
    {
        "teacherId": "edureka-expert",
        "teacherName": "Edureka Expert",
        "channel": "Edureka",
        "domain": "AI and ML fundamentals",
        "summary": "Builds the ladder from AI to ML to deep learning before leading into Generative AI, which makes the field less intimidating for new learners.",
        "teachingStyle": "Structured, beginner-friendly, concept ladder from basics to advanced",
        "signatureTopics": ["AI hierarchy", "learning paradigms", "NLP basics", "deep learning"],
        "teacherVoice": "AI is about making computers smarter and more helpful in everyday life.",
        "realWorldImpact": "Useful for learners who need the big picture before specializing in GenAI.",
    },
    {
        "teacherId": "andrew-brown",
        "teacherName": "Andrew Brown",
        "channel": "ExamPro / Andrew Brown",
        "domain": "GenAI development roadmap",
        "summary": "Connects concepts to real tool choices, maturity stages, deployment, and the broader ecosystem around AI products.",
        "teachingStyle": "Roadmap-driven, implementation focused, cloud-aware and pragmatic",
        "signatureTopics": ["maturity models", "modalities", "evaluation", "deployment", "tooling"],
        "teacherVoice": "Broad and practical GenAI knowledge gives you the flexibility to move in any technical direction.",
        "realWorldImpact": "Great for learners turning theory into projects and stack decisions.",
    },
    {
        "teacherId": "simplilearn-specialist",
        "teacherName": "Simplilearn Specialist",
        "channel": "Simplilearn",
        "domain": "Prompting, tools, business use, and agents",
        "summary": "Frames GenAI as a modern workforce skill, then shows how prompts, workflows, and agents create business value.",
        "teachingStyle": "Product-minded, workforce-oriented, use-case heavy",
        "signatureTopics": ["prompt anatomy", "AI agents", "multimodal AI", "productivity", "business use cases"],
        "teacherVoice": "Generative AI is no longer optional; it is becoming a must-know capability of the modern workforce.",
        "realWorldImpact": "Strong fit for professionals building useful day-to-day workflows.",
    },
    {
        "teacherId": "stanford-cs229",
        "teacherName": "Stanford CS229",
        "channel": "Stanford CS229",
        "domain": "Training science and alignment",
        "summary": "Explains what really matters during model building: data, scaling, alignment, optimization, and systems constraints.",
        "teachingStyle": "Rigorous, research-oriented, systems-and-training depth",
        "signatureTopics": ["pretraining", "scaling laws", "SFT", "RLHF", "data quality"],
        "teacherVoice": "Most people talk about architecture, but in practice data, evaluation, and systems matter just as much.",
        "realWorldImpact": "Best for learners who want to understand why modern models behave the way they do.",
    },
]


TRACK_TEACHER_MAP = {
    "AI Foundations": "edureka-expert",
    "Machine Learning Basics": "edureka-expert",
    "Core ML Algorithms": "edureka-expert",
    "Deep Learning and Neural Networks": "stanford-cs229",
    "Generative AI Fundamentals": "simplilearn-specialist",
    "Large Language Models": "gaurav-sen",
    "Training and Alignment": "stanford-cs229",
    "Prompt Engineering": "simplilearn-specialist",
    "RAG and Context Engineering": "gaurav-sen",
    "AI Agents and Agentic AI": "simplilearn-specialist",
    "Multimodal AI": "andrew-brown",
    "Responsible AI and Safety": "stanford-cs229",
    "Building AI Applications": "andrew-brown",
    "MLOps and LLMOps": "andrew-brown",
    "AI Tools and Ecosystem": "andrew-brown",
    "Business Applications": "simplilearn-specialist",
    "Career and Skills": "andrew-brown",
    "Hands-on Practice": "andrew-brown",
    "Advanced GenAI Concepts": "gaurav-sen",
    "Data and Evaluation": "stanford-cs229",
}


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def read_paragraphs(path: Path) -> list[str]:
    return [paragraph.text.strip() for paragraph in Document(path).paragraphs if paragraph.text.strip()]


def parse_100_page_workbook() -> tuple[list[dict], list[dict]]:
    paragraphs = read_paragraphs(DOC_100)
    heading_pattern = re.compile(r"^Page (\d{3}) - (.+)$")
    lesson_starts = [index for index, line in enumerate(paragraphs) if heading_pattern.match(line)]

    lessons = []
    for position, start in enumerate(lesson_starts):
        end = lesson_starts[position + 1] if position + 1 < len(lesson_starts) else len(paragraphs)
        block = paragraphs[start:end]
        match = heading_pattern.match(block[0])
        if not match:
            continue
        page = match.group(1)
        title = match.group(2)
        track = block[1] if len(block) > 1 else "Workbook"
        content = block[2] if len(block) > 2 else ""
        thinking = ""
        recap = ""
        extras = []
        for line in block[3:]:
            if line.startswith("How to think about it:"):
                thinking = line.replace("How to think about it:", "", 1).strip()
            elif line.startswith("Mini recap:"):
                recap = line.replace("Mini recap:", "", 1).strip()
            elif not line.startswith("Glossary Part"):
                extras.append(line)
        lessons.append(
            {
                "page": page,
                "title": title,
                "track": track,
                "slug": slugify(title),
                "teacherId": TRACK_TEACHER_MAP.get(track, "andrew-brown"),
                "content": content,
                "thinking": thinking,
                "recap": recap,
                "extras": extras,
            }
        )

    grouped = OrderedDict()
    for lesson in lessons:
        grouped.setdefault(lesson["track"], []).append(lesson)

    tracks = []
    for index, (track_title, track_lessons) in enumerate(grouped.items(), start=1):
        tracks.append(
            {
                "id": slugify(track_title),
                "order": index,
                "title": track_title,
                "teacherId": TRACK_TEACHER_MAP.get(track_title, "andrew-brown"),
                "pageRange": f"{track_lessons[0]['page']}–{track_lessons[-1]['page']}",
                "summary": track_lessons[0]["content"],
                "lessonCount": len(track_lessons),
            }
        )

    return tracks, lessons


def parse_detailed_chapters() -> list[dict]:
    paragraphs = read_paragraphs(DOC_DETAILED)
    chapter_pattern = re.compile(r"^Chapter (\d+): (.+)$")
    chapters = []
    current = None
    current_section = None
    mode = "sections"

    for line in paragraphs[2:]:
        chapter_match = chapter_pattern.match(line)
        if chapter_match:
            if current:
                if current_section:
                    current["sections"].append(current_section)
                chapters.append(current)
            current = {
                "id": f"chapter-{chapter_match.group(1)}",
                "number": int(chapter_match.group(1)),
                "title": chapter_match.group(2),
                "sections": [],
                "exercises": [],
                "recap": "",
            }
            current_section = None
            mode = "sections"
            continue

        if not current:
            continue

        if line == "Exercises & Practice Tasks":
            if current_section:
                current["sections"].append(current_section)
                current_section = None
            mode = "exercises"
            continue

        if line == "Mini Recap":
            if current_section:
                current["sections"].append(current_section)
                current_section = None
            mode = "recap"
            continue

        if mode == "sections":
            if line.startswith("Beginner Explanation:"):
                if current_section:
                    current_section["beginnerTip"] = line.replace("Beginner Explanation:", "", 1).strip()
                continue
            if current_section and not current_section.get("summary"):
                current_section["summary"] = line
            else:
                if current_section:
                    current["sections"].append(current_section)
                current_section = {"heading": line, "summary": "", "beginnerTip": ""}
        elif mode == "exercises":
            current["exercises"].append(line)
        elif mode == "recap":
            current["recap"] = line

    if current:
        if current_section:
            current["sections"].append(current_section)
        chapters.append(current)

    return chapters


def parse_revision_topics() -> list[dict]:
    paragraphs = read_paragraphs(DOC_REVISION)
    topic_pattern = re.compile(r"^(\d+)\. (.+)$")
    topics = []
    current = None
    for line in paragraphs[2:]:
        match = topic_pattern.match(line)
        if match:
            if current:
                topics.append(current)
            current = {
                "number": int(match.group(1)),
                "title": match.group(2),
                "points": [],
                "beginnerTip": "",
            }
            continue

        if not current:
            continue

        if line.startswith("Beginner Tip:"):
            current["beginnerTip"] = line.replace("Beginner Tip:", "", 1).strip()
        else:
            current["points"].append(line)

    if current:
        topics.append(current)

    return topics


def build_site_json() -> dict:
    tracks, lessons = parse_100_page_workbook()
    chapters = parse_detailed_chapters()
    revision_topics = parse_revision_topics()

    return {
        "site": {
            "name": "PromptCraft Academy",
            "tagline": "A GitHub-ready, zero-cost multi-page GenAI learning experience",
            "description": "This site consolidates three beginner workbooks into an approachable, interactive website that teaches AI, machine learning, and generative AI through guided tracks, chapter notes, revision pages, teacher-led perspectives, and hands-on practice.",
        },
        "stats": {
            "lessons": len(lessons),
            "tracks": len(tracks),
            "chapters": len(chapters),
            "revisionTopics": len(revision_topics),
            "teachers": len(TEACHERS),
        },
        "highlights": [
            "100 bite-sized workbook lessons grouped into 20 topic tracks",
            "5 detailed beginner chapters with practice exercises",
            "15 quick-revision topic summaries for fast review",
            "5 teacher perspectives to make the content feel more dynamic",
            "Static hosting friendly architecture for GitHub Pages",
        ],
        "tracks": tracks,
        "lessons": lessons,
        "chapters": chapters,
        "revisionTopics": revision_topics,
        "teachers": TEACHERS,
        "teacherLessons": [
            {
                "teacherId": lesson["teacherId"],
                "page": lesson["page"],
                "moduleTitle": lesson["title"],
                "difficulty": "Beginner" if int(lesson["page"]) <= 40 else "Intermediate",
                "summary": lesson["content"],
            }
            for lesson in lessons
        ],
        "sources": [
            "Detailed_Beginner_AI_Workbook.docx",
            "AI_Generative_AI_100_Page_Beginner_Workbook.docx",
            "Generative_AI_Workbook.docx",
        ],
    }


def write_json(data: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    SITE_JSON.write_text(json.dumps(data, indent=2, ensure_ascii=True), encoding="utf-8")


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="DDF4EC")
    font = Font(bold=True, color="0D5F4D")
    for cell in row:
        cell.fill = fill
        cell.font = font


def build_workbook(data: dict) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    def create_sheet(name: str, headers: list[str], rows: list[list[str]]) -> None:
        sheet = workbook.create_sheet(title=name)
        sheet.append(headers)
        style_header(sheet[1])
        for row in rows:
            sheet.append(row)
        for column in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 14), 48)

    create_sheet(
        "Teachers",
        ["teacherId", "teacherName", "channel", "domain", "teachingStyle", "teacherVoice", "realWorldImpact"],
        [
            [
                teacher["teacherId"],
                teacher["teacherName"],
                teacher["channel"],
                teacher["domain"],
                teacher["teachingStyle"],
                teacher["teacherVoice"],
                teacher["realWorldImpact"],
            ]
            for teacher in data["teachers"]
        ],
    )

    create_sheet(
        "LessonCatalog",
        ["page", "title", "track", "teacherId", "summary"],
        [
            [lesson["page"], lesson["title"], lesson["track"], lesson["teacherId"], lesson["content"]]
            for lesson in data["lessons"]
        ],
    )

    create_sheet(
        "Users",
        ["userId", "name", "email", "passwordHash", "createdAt", "updatedAt", "status", "source"],
        [],
    )
    create_sheet(
        "Sessions",
        ["sessionId", "userId", "email", "event", "timestamp", "branch", "page"],
        [],
    )
    create_sheet(
        "Progress",
        ["entryId", "userId", "email", "moduleId", "moduleTitle", "score", "xp", "completed", "timestamp"],
        [],
    )
    create_sheet(
        "Activity",
        ["activityId", "userId", "email", "type", "title", "details", "timestamp"],
        [],
    )
    create_sheet(
        "Meta",
        ["key", "value"],
        [
            ["repoOwner", "neonj3"],
            ["repoName", "PromptCraftAcademy"],
            ["defaultBranch", "testing-features"],
            ["workbookPath", "data/promptcraft-github-database.xlsx"],
            ["notes", "This workbook is designed to be updated through the GitHub Contents API from a static frontend."],
        ],
    )

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(DB_WORKBOOK)


def main() -> None:
    data = build_site_json()
    write_json(data)
    build_workbook(data)
    print(f"Wrote {SITE_JSON}")
    print(f"Wrote {DB_WORKBOOK}")


if __name__ == "__main__":
    main()
